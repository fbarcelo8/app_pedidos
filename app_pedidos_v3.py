import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
from PIL import Image
from modules.login import load_users, login_user, logout_user

# Función para extraer el código de almacén del DataFrame unificado
def extraer_codigo_almacen(df_unificado):
    try:
        codigo_almacen = df_unificado['Cód. almacén'].iloc[0]
    except (IndexError, KeyError):
        codigo_almacen = "codigo_no_encontrado"
    return codigo_almacen

def procesar_producto(df_lineas):
    df_producto_filtrado = df_lineas[df_lineas['Bulk?'] == 0]
    df_producto_agrupado = df_producto_filtrado.groupby(['Nº', 'Descripción'], as_index=False).agg({'Cantidad': 'sum'})
    return df_producto_agrupado

def procesar_empleados(df_lineas):
    df_empleados_filtrado = df_lineas[df_lineas['Bulk?'] == 0]
    df_empleados_unicos = df_empleados_filtrado[['Empleado']].drop_duplicates().copy()
    df_empleados_unicos['Empleado '] = df_empleados_unicos['Empleado']
    return df_empleados_unicos

def procesar_hoja_empleado_por_aeropuerto(df_lineas):
    df_empleados_filtrado = df_lineas[df_lineas['Bulk?'] == 0].copy()
    aeropuertos = df_empleados_filtrado['Aeropuerto'].unique()
    tablas_por_aeropuerto = {}

    for aeropuerto in aeropuertos:
        df_aeropuerto = df_empleados_filtrado[df_empleados_filtrado['Aeropuerto'] == aeropuerto].copy()
        tablas_por_aeropuerto[aeropuerto] = df_aeropuerto[['Nº', 'Descripción', 'Empleado', 'Cantidad', 'Aeropuerto']]

    return tablas_por_aeropuerto

def procesar_pedido(pedido_df):
    try:
        aeropuerto = pedido_df['Descripción'].str.extract(r'Preparacion Pedido:(\w+)').dropna().iloc[0, 0]
    except (IndexError, KeyError):
        print("No se encontró 'Preparacion Pedido' en este archivo o falta la columna 'Descripción'.")
        return pedido_df
    
    pedido_df['Aeropuerto'] = aeropuerto
    pedido_filtrado = pedido_df[~pedido_df['Nº'].str.contains('PREPARACION|PICKING', case=False, na=False)].copy()

    if 'Empleado' in pedido_filtrado.columns:
        pedido_filtrado.loc[:, 'Empleado'] = pedido_filtrado['Empleado'].fillna(aeropuerto)
        pedido_filtrado.loc[:, 'Bulk?'] = pedido_filtrado['Empleado'].apply(lambda x: 1 if 'BULK' in str(x).upper() else 0)
    
    # Verificar si existe la columna 'Talla'
    if 'Talla' not in pedido_filtrado.columns:
        # Extraer la información de la talla desde la columna 'Descripción'
        pedido_filtrado['Talla'] = pedido_filtrado['Descripción'].str.extract(r'T\.(\w+)', expand=False)
        
        # Si no se encuentra 'T.', buscar un número de dos dígitos
        pedido_filtrado['Talla'] = pedido_filtrado['Talla'].fillna(
            pedido_filtrado['Descripción'].str.extract(r'(\b\d{2}\b)', expand=False)
        )
    
    return pedido_filtrado

def procesar_pedidos(lista_archivos):
    pedidos_unificados = []
    for archivo in lista_archivos:
        try:
            pedido_df = pd.read_excel(archivo)
        except Exception as e:
            print(f"No se pudo cargar el archivo: {e}")
            continue
        pedido_procesado = procesar_pedido(pedido_df)
        pedidos_unificados.append(pedido_procesado)

    if pedidos_unificados:
        df_unificado = pd.concat(pedidos_unificados, ignore_index=True)
    else:
        df_unificado = pd.DataFrame()
    
    return df_unificado

def aplicar_estilo_encabezado_columnas(ws, fila, num_columnas, color_fondo='7A9CCF'):
    color_fondo_encabezado = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")
    fuente_encabezado = Font(bold=True)
    alineacion = Alignment(horizontal="center", vertical="center")
    for col in range(1, num_columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = color_fondo_encabezado
        celda.font = fuente_encabezado
        celda.alignment = alineacion

def ajustar_ancho_columnas(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

def aplicar_formato_tabla(ws, inicio_fila, fin_fila, num_columnas, nombre_tabla):
    rango_tabla = f"A{inicio_fila}:{get_column_letter(num_columnas)}{fin_fila}"
    tabla = Table(displayName=nombre_tabla, ref=rango_tabla)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

def aplicar_estilo_filas(ws, inicio_fila, fin_fila):
    color1 = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    color2 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    for row in ws.iter_rows(min_row=inicio_fila, max_row=fin_fila, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = color1 if cell.row % 2 == 0 else color2

def crear_archivo_excel_en_memoria(lista_archivos):
    df_lineas = procesar_pedidos(lista_archivos)
    if df_lineas.empty:
        print("No se pudo procesar ningún pedido, el archivo Excel no se creará.")
        return None, None
    
    codigo_almacen = extraer_codigo_almacen(df_lineas)
    df_producto = procesar_producto(df_lineas)
    df_empleados = procesar_empleados(df_lineas)
    
    tablas_por_aeropuerto = procesar_hoja_empleado_por_aeropuerto(df_lineas)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_producto.to_excel(writer, sheet_name='Producto', index=False)
        df_empleados.to_excel(writer, sheet_name='Empleados', index=False)
        
        # Formatear la hoja 'Empleado'
        wb = writer.book
        ws_empleado = wb.create_sheet(title='Empleado')
        fila_actual = 1
        contador_tablas = 1

        # Insertar cada tabla por aeropuerto con su encabezado
        for aeropuerto, df_aeropuerto in tablas_por_aeropuerto.items():
            if fila_actual > 1:
                fila_actual += 2

            # Insertar título del aeropuerto
            columnas_totales = len(df_aeropuerto.columns)
            ws_empleado.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=columnas_totales)
            celda_titulo = ws_empleado.cell(row=fila_actual, column=1, value=f"Aeropuerto: {aeropuerto}")
            celda_titulo.alignment = Alignment(horizontal='center')
            aplicar_estilo_encabezado_columnas(ws_empleado, fila_actual, columnas_totales, '7A9CCF')
            fila_actual += 1

            # Insertar encabezados de columnas
            for col_idx, columna in enumerate(df_aeropuerto.columns, start=1):
                ws_empleado.cell(row=fila_actual, column=col_idx, value=columna)
            aplicar_estilo_encabezado_columnas(ws_empleado, fila_actual, len(df_aeropuerto.columns))
            inicio_tabla = fila_actual
            fila_actual += 1

            # Insertar datos de la tabla
            for fila in df_aeropuerto.itertuples(index=False):
                for col_idx, valor in enumerate(fila, start=1):
                    ws_empleado.cell(row=fila_actual, column=col_idx, value=valor)
                fila_actual += 1
            
            # Aplicar formato de tabla para el rango del aeropuerto
            fin_tabla = fila_actual - 1
            aplicar_formato_tabla(ws_empleado, inicio_tabla, fin_tabla, len(df_aeropuerto.columns), f"Tabla_{contador_tablas}")
            contador_tablas += 1
            ajustar_ancho_columnas(ws_empleado)

        df_lineas.to_excel(writer, sheet_name='Líneas', index=False)

        # Aplicar formato y estilo a las otras hojas
        ws_producto = wb['Producto']
        aplicar_estilo_encabezado_columnas(ws_producto, 1, len(df_producto.columns), '7A9CCF')
        aplicar_formato_tabla(ws_producto, 1, len(df_producto) + 1, len(df_producto.columns), "Tabla_Producto")
        aplicar_estilo_filas(ws_producto, 2, len(df_producto) + 1)
        ajustar_ancho_columnas(ws_producto)

        ws_empleados = wb['Empleados']
        aplicar_estilo_encabezado_columnas(ws_empleados, 1, len(df_empleados.columns), '7A9CCF')
        aplicar_formato_tabla(ws_empleados, 1, len(df_empleados) + 1, len(df_empleados.columns), "Tabla_Empleados")
        aplicar_estilo_filas(ws_empleados, 2, len(df_empleados) + 1)
        ajustar_ancho_columnas(ws_empleados)

        ws_lineas = wb['Líneas']
        aplicar_estilo_encabezado_columnas(ws_lineas, 1, len(df_lineas.columns), '7A9CCF')
        aplicar_formato_tabla(ws_lineas, 1, len(df_lineas) + 1, len(df_lineas.columns), "Tabla_Lineas")
        aplicar_estilo_filas(ws_lineas, 2, len(df_lineas) + 1)
        ajustar_ancho_columnas(ws_lineas)

    output.seek(0)
    return output, codigo_almacen

def aplicar_imagenes_y_formato():
    im = Image.open('resources/delagencia_sl_logo.jpg')

    st.set_page_config(
        page_title='Procesamiento de Pedidos',
        page_icon=im
    )

    col1, col2, col3 = st.columns([1, 6, 1])
        
    with col1:
        st.image('resources/logo.jpg', width=170)

    with col3:
        st.image('resources/delagencia_sl_logo.jpg', width=90)

def main():
    if 'user_state' not in st.session_state:
        st.session_state.user_state = {
            'username': '',
            'password': '',
            'logged_in': False
        }

    # Inicializa la clave `file_uploader_key`
    if 'file_uploader_key' not in st.session_state:
        st.session_state.file_uploader_key = 0  # Clave para reiniciar el file_uploader

    users = load_users()

    # Login Section
    if not st.session_state.user_state['logged_in']:
        aplicar_imagenes_y_formato()
        st.markdown('## Procesamiento de Pedidos')

        username = st.text_input('Usuario')
        password = st.text_input('Contraseña', type='password')
        submit = st.button('Iniciar sesión')

        if submit:
            user_info = users.get(username)
            if user_info and user_info["password"] == password:
                st.session_state.user_state['logged_in'] = True
                st.session_state.user_state['username'] = username
                st.session_state.user_state['name'] = user_info['name']
                st.rerun()  # Recarga la página después del login
            else:
                st.warning('Usuario / contraseña incorrecto')

    # Main Content for Logged-in Users
    else:
        aplicar_imagenes_y_formato()
        st.title("Procesamiento de Pedidos")
        st.success(f"¡Bienvenido {st.session_state.user_state['name']}!")

        # Cargar archivos usando `file_uploader_key`
        uploaded_files = st.file_uploader("Cargar archivos Excel", type="xlsx", accept_multiple_files=True, key=st.session_state.file_uploader_key)

        # Mostrar botón "Eliminar todos" solo si hay archivos cargados
        if uploaded_files:
            col1, col2 = st.columns([0.8, 0.2])
            with col2:
                if st.button("Eliminar todos"):
                    # Reiniciar el file_uploader al actualizar la clave
                    st.session_state.file_uploader_key += 1
                    st.rerun()  # Forzar la recarga después de eliminar todos los archivos

        # Mostrar el botón "Procesar archivos" solo si hay archivos cargados
        if uploaded_files:
            if st.button("Procesar archivos"):
                with st.spinner('Procesando...'):
                    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                    output, codigo_almacen = crear_archivo_excel_en_memoria(uploaded_files)
                    if output:
                        st.success("Archivos procesados correctamente")
                        st.download_button(
                            label="Descargar archivo Excel",
                            data=output.getvalue(),
                            file_name=f"pedidos_{codigo_almacen}_{fecha_hoy}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error("Error al procesar los archivos.")

        if st.button('Cerrar sesión'):
            logout_user()
            st.session_state.user_state['logged_in'] = False
            st.session_state.file_uploader_key += 1  # Resetea el uploader al cerrar sesión
            st.rerun()  # Forzar la recarga después de cerrar sesión

if __name__ == "__main__":
    main()
